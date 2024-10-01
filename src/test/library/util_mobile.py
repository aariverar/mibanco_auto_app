from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
import os
import time
from openpyxl import Workbook, load_workbook


def capture_screenshot(xdriver, filename, directory='screenshots'):
    if not os.path.exists(directory):
        os.makedirs(directory)
    filepath = os.path.join(directory, filename)
    xdriver.save_screenshot(filepath)
            

def save_excel_locally(file_path, row_data):
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "SCENARIO", "FECHA", "HORA", "TIMER", "ESTADO", "FRAMEWORK", "HOSTNAME","DEVICE_NAME","UID","APP"])
    else:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    
    sheet.append(row_data)
    workbook.save(filename=file_path)

def get_next_id(file_path):
    if not os.path.exists(file_path):
        return 1  # Si el archivo no existe, empezamos desde el ID 1
    
    workbook = load_workbook(file_path)
    sheet = workbook.active
    return sheet.max_row  # El próximo ID es el número total de filas + 1

def scrollMobile(mdriver):
    mdriver.execute_script("mobile: shell", {
                    'command': 'input', 
                    'args': ['swipe', '500', '1000', '500', '500']  # Ajusta estas coordenadas según el tamaño de tu dispositivo
                })
        